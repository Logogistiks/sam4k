from __future__ import annotations

__all__ = ["COM_CODES", "CODE_STX", "CODE_ENQ", "CODE_ACK", "CODE_CR", "CODE_NAK", "CODE_ETB", "CODE_EXIT", "CODE_BAR", "CODE_NOBAR", "Shot", "Transmission", "checksum_xor", "open_file", "save_data"]
__version__ = "1.2.3"
__author__ = "Jan Seifert <sam4k@logogistiks.de>"

#built in modules
import os
import re
from copy import deepcopy
from datetime import datetime
from math import trunc
from time import sleep
from dataclasses import dataclass

#external modules
try:
    import openpyxl
    import openpyxl.cell
    import openpyxl.styles
    from serial import EIGHTBITS, PARITY_NONE, STOPBITS_ONE, Serial
    from serial.tools import list_ports
    import beaupy
    from colorama import init, Fore, Back, Style
    from pynput import keyboard
except ImportError as e:
    print(f"Fehler beim importieren: {e}.\n Bitte die erforderlichen Module mit 'pip install -r requirements.txt' installieren.")
    input("Drücke Enter zum Beenden...")
    raise SystemExit

init(convert=True) # colorama init for Windows compatibility

#################### Begin Basic Settings ####################

PORT = {"nt": "COM3", "posix": "/dev/ttyUSB0"}[os.name]
"""The serial port of the SAM4000 device"""

SHOTS_PER_SERIES = 10 # should be 1, 2, 5, or a multiple of 10
"""How many shots should be saved in a series (one row in the excel file)"""

pattern_header = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # light blue
pattern_mark1 = openpyxl.styles.PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid") # light yellow
pattern_mark2 = openpyxl.styles.PatternFill(start_color="F08080", end_color="F08080", fill_type="solid") # light coral

LOG_TRANSMISSIONS = False
"""Whether to log the raw bytes received from the SAM4000 device to a file"""

CHSUM_RETRY = 3

#################### Begin Program Logic #####################

COM_CODES = [
    CODE_STX := b"\x02",
    CODE_ENQ := b"\x05",
    CODE_ACK := b"\x06",
    CODE_CR := b"\x0D",
    CODE_NAK := b"\x15",
    CODE_ETB := b"\x17",
    CODE_EXIT := b"\xB0",
    CODE_BAR := b"\xB1",
    CODE_NOBAR := b"\xB2"
]
"""Codes for communication with the SAM4000 device"""

@dataclass
class Shot:
    """Represents a shot in the transmission"""
    ring: float | None = None
    div: float | None = None
    x: int | None = None
    y: int | None = None

    def __str__(self) -> str:
        """Returns a human readable representation of the Shot object"""
        return f"Shot(ring={self.ring}, div={self.div}, x={self.x}, y={self.y})"

class Transmission:
    "This class implements handling of a typical transmission by the SAM4000 device, which is received in bytes via serial connection."

    def __init__(self, barcode: str=None, manual_code: str=None, target_type: str=None, target_num: int=None, div: float=None, shots_num: int=None, shots: list[Shot]=None) -> None:
        """Initializes a Transmission object with the given parameters, allthough usage of `Transmission.create_empty` or `Transmission.from_bytes` is recommended."""
        self.barcode: str = barcode
        self.manual_code: str = manual_code
        self.target_type: str = target_type
        self.target_num: int = target_num
        self.div: float = div
        self.shots_num: int = shots_num
        self.shots: list[Shot] = shots

    @staticmethod
    def create_empty() -> Transmission:
        """returns an empty Transmission object with attributes of respective type instead of `None`"""
        return Transmission(barcode="", manual_code="", target_type="", target_num=0, div=0.0, shots_num=0, shots=[])

    @staticmethod
    def example() -> Transmission:
        """returns an example Transmission object with filled in dummy data"""
        return Transmission(barcode="01236789", manual_code="98763210", target_type="LG", target_num=10, div=1.0, shots_num=10, shots=[
            Shot(ring=8.6, div=532.7, x=-236, y=508),
            Shot(ring=9.0, div=500.0, x=-200, y=500),
            Shot(ring=4.9, div=276.0, x=-154, y=525),
            Shot(ring=0.0, div=None, x=None, y=None), # missed shot
            Shot(ring=10.0, div=0.0, x=-111, y=500),
            Shot(ring=9.5, div=None, x=-100, y=500), # manually corrected shot
            Shot(ring=8.1, div=396.0, x=-52, y=-156),
            Shot(ring=7.0, div=400.0, x=-50, y=-200),
            Shot(ring=6.5, div=350.0, x=-25, y=-250),
            Shot(ring=5.0, div=300.0, x=-10, y=-300)])

    def __str__(self) -> str:
        """Returns a human readable representation of the Transmission object"""
        res = ""
        res += f"Transmission(\n"
        res += f"    barcode={self.barcode},\n"
        res += f"    manual_code={self.manual_code},\n"
        res += f"    target_type={'None' if self.target_type is None else f'{self.target_type}'},\n"
        res += f"    target_num={self.target_num},\n"
        res += f"    div={self.div},\n"
        res += f"    shots_num={self.shots_num},\n"
        res += f"    shots=[\n"
        for i, shot in enumerate(self.shots):
            res += f"        {shot}{',' if i != len(self.shots)-1 else ''}\n"
        res += f"    ]\n"
        res += f")"
        return res

    @staticmethod
    def _valid_barcode(bc: str) -> bool:
        """Checks if a barcode string is of valid form"""
        return bool(re.fullmatch(r"[0-9]{8}", bc))

    @staticmethod
    def _valid_manual_code(mc: str) -> bool:
        """Checks if a manual code string is of valid form"""
        return bool(re.fullmatch(r"[0-9]{8}", mc))

    @staticmethod
    def _valid_target_type(tt: str) -> bool:
        """Checks if a target type string is of valid form"""
        return tt in ("LG", "LP", "KK", "ZS", "LS")

    @staticmethod
    def _valid_target_num(tn: str) -> bool:
        """Checks if a target number string is of valid form"""
        return bool(re.fullmatch(r"[0-9]{2}", tn))

    @staticmethod
    def _valid_div(div: str) -> bool:
        """Checks if a division factor string is of valid form"""
        return bool(re.fullmatch(r"[0-9]\.[0-9]", div))

    @staticmethod
    def _valid_shot_number(sn: str) -> bool:
        """Checks if a shot number string is of valid form"""
        return bool(re.fullmatch(r"[0-9]{2}", sn))

    @staticmethod
    def from_bytes(byt: bytes, log: bool=False) -> Transmission:
        """Parses the given bytes into a Transmission object and returns it."""
        if log:
            print(byt)
        trans = Transmission.create_empty()
        bc, mc, tt, tn, div, sn, *s = [part.decode("unicode-escape") for part in byt.split(CODE_CR)] # remove last empty string
        s = [item for item in s if item] # remove empty strings
        if log:
            for item in [bc, mc, tt, tn, div, sn, s]:
                print(item)
        if len(s) % 4 != 0: # s is a list of strings, each 4 strings represent a shot
            raise ValueError("bytes are of invalid form, shot data does not make sense (not a multiple of 4)")
        # technically the ? check is not necessary, but is left for clarity
        if not "?" in bc and Transmission._valid_barcode(bc):
            trans.barcode = bc
        if not "?" in mc and Transmission._valid_manual_code(mc):
            trans.manual_code = mc
        if not "?" in tt and Transmission._valid_target_type(tt):
            trans.target_type = tt
        if not "?" in tn and Transmission._valid_target_num(tn):
            trans.target_num = int(tn)
        if not "?" in div and Transmission._valid_div(div):
            trans.div = float(div)
        if not "?" in sn and Transmission._valid_shot_number(sn):
            trans.shots_num = int(sn)
        trans.shots = []
        for i in range(0, len(s), 4):
            trans.shots.append(Shot(
                ring=float(s[i]) if not "?" in s[i] else None,
                div=float(s[i+1]) if not "?" in s[i+1] else None,
                x=int(s[i+2]) if not "?" in s[i+2] else None,
                y=int(s[i+3]) if not "?" in s[i+3] else None
            ))
        #*Note: ↓ maybe useful later for distinguishing between cases: ↓
        #   ring is 0 and div is ? => missed shot
        #   ring > 0 und div is ? => manually corrected shot
        #   rind > 0 und Div > 0 => normal shot
        return trans

    @staticmethod
    def receive(ser: Serial, retry_infinite: bool=False) -> Transmission | None:
        """Receives bytes from the given serial port and returns it as a Transmission object. \\
        If checksum is wrong and max retries is reached, sends CODE_ACK and returns None."""
        retries = 0
        while True:
            response = ser.read_until(b"\x24")[:-1] # read until dollar sign exclusively
            if LOG_TRANSMISSIONS:
                log(response)
            data, checksum = response.split(CODE_ETB)
            calc_checksum = checksum_xor(CODE_STX + data + CODE_ETB)
            if calc_checksum != ord(checksum):
                if retries <= CHSUM_RETRY or retry_infinite: # retry_infinite overrides CHSUM_RETRY
                    ser.write(CODE_NAK)
                    retries += 1
                    continue
                else: # can only happen if retry_infinite is False
                    ser.write(CODE_ACK) # "pretending"
                    return None # None is the flag for a failed transmission
            else:
                break
        return Transmission.from_bytes(data)

    def get_valid_shot_num(self) -> int:
        """Returns the number of valid shots in the transmission"""
        return sum(1 for shot in self.shots if shot.ring is not None)

    def get_invalid_shot_num(self) -> int:
        """Returns the number of invalid shots in the transmission"""
        return sum(1 for shot in self.shots if shot.ring is None)

    def get_manual_corrected_num(self) -> int:
        """Returns the number of shots that were manually corrected"""
        return sum(1 for shot in self.shots if shot.ring is not None and shot.div is None)

    def get_valid_shots(self, fill: int=None) -> list[Shot]:
        """Returns a list of valid shots in the transmission. \\
        If `fill` is given, pads the list with empty shots to the given length."""
        valid_shots = [shot for shot in self.shots if shot.ring is not None]
        if fill is not None and len(valid_shots) < fill:
            valid_shots += [Shot(0.0, None, None, None) for _ in range(fill - len(valid_shots))]
        return valid_shots

class MemoryHandler:
    """This class implements a dynamic memory for shot data, both short-term and long-term for multiple people."""
    def __init__(self, shots_per_strip: int) -> None:
        """Initializes an empty Memory object"""
        self.MEM_short: list[Shot] = [] # short-term memory, used for the current series until full
        self.MEM_long: dict[str, list[list[Shot]]] = {} # long-term memory, used for all series of all people, key is the person's name
        self.person_count = 0 # number of people processed so far
        self.strip_count = 0 # number of strips processed for current person
        self.current_person: str = None
        self.SHOTS_PER_STRIP = shots_per_strip

    def update_person(self, name: str) -> None:
        """Updates the current person with the given name."""
        self.current_person = name
        self.person_count += 1
        self.strip_count = 0 # reset strip count for new person
        if name not in self.MEM_long: # should always be the case
            self.MEM_long[name] = [] # initialize long-term memory for this person

    def update_memory(self, trans: Transmission) -> None:
        """Updates the memory with the given transmission data. \\
        New data will be added to short term memory, and if it is full, it will be moved to long term memory."""

        # update short term memory with the current transmission
        # fill with empty shots if less than expected, discard surplus if more than expected
        self.MEM_short += trans.get_valid_shots(fill=self.SHOTS_PER_STRIP)[:self.SHOTS_PER_STRIP] # looks hacky, but works. Tested all cases > < ==
        self.strip_count += 1 # strip processed

        # check if short term can be carried over to long term
        if len(self.MEM_short) >= SHOTS_PER_SERIES:
            self.MEM_long[self.current_person].append(deepcopy(self.MEM_short[:SHOTS_PER_SERIES])) # discard surplus, shouldnt happen though
            self.MEM_short.clear() # clear short-term memory for next strip

    def finalize(self) -> None:
        """Prepares the memory for saving to file."""

        # remove empty people, can happen if pressed Enter immediately after entering a name
        self.MEM_long = {k: v for k, v in self.MEM_long.items() if v}

def log(content: str | bytes) -> None:
    """Logs the given content (readableBuffer) to a file in the log directory, filename is current time"""
    if not os.path.exists("log"):
        os.mkdir("log")
    with open(os.path.join("log", f"log-{nowtime()}.bin"), "wb" if isinstance(content, bytes) else "w") as f:
        f.write(content)

def clear() -> None:
    """Clears the console"""
    os.system(("cls" if os.name == "nt" else "clear"))

def nowtime(pretty: bool=False) -> str:
    """Returns the current time in YYYY_MM_DD-HH_MM_SS format. \\
    If `pretty` is True, returns a human readable format."""
    if pretty:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return datetime.now().strftime("%Y_%m_%d-%H_%M_%S")

def checksum_xor(byt: bytes) -> int:
    """Calculates the XOR checksum of the given bytes. \\
    Works by XORing all the bytes together."""
    chsum = 0
    for b in byt:
        chsum ^= b
    return chsum

def record_keypresses(t: float=1) -> list[keyboard.Key | keyboard.KeyCode]:
    """Records all keyboard activity in the next `t` seconds synchronously, blocking the program flow."""
    pressed = []
    listener = keyboard.Listener(on_press=None, on_release=pressed.append) # looks a bit cursed
    listener.start()
    sleep(t)
    listener.stop()
    return pressed

def open_file(fname: str) -> None:
    """Opens the file with the default program"""
    if os.name == "nt":
        os.startfile(fname, "open")
    elif os.name == "posix":
        try:
            os.system(f"xdg-open {fname}")
        except:
            print("Could not open the file")
    else:
        print("Could not open the file")

def set_cell(ws, row: int, col: int, value=None, fill=None, b_left: bool=False, b_right: bool=False, b_top: bool=False, b_bottom: bool=False, center_h: bool=False, center_v: bool=False) -> None:
    """Sets the value of a cell and applies the given fill and border settings"""
    cell: openpyxl.cell.Cell = ws.cell(row, col)
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if any((b_left, b_right, b_top, b_bottom)):
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin") if b_left else None,
            right=openpyxl.styles.Side(style="thin") if b_right else None,
            top=openpyxl.styles.Side(style="thin") if b_top else None,
            bottom=openpyxl.styles.Side(style="thin") if b_bottom else None)
    if any((center_h, center_v)):
        cell.alignment = openpyxl.styles.Alignment(horizontal="center" if center_h else None, vertical="center" if center_v else None)

def draw_header(ws, start_cell: tuple[int]=(1, 1)) -> tuple[int]:
    """Draws the header on the excel worksheet. Returns the bottom right cell of the header. \\
    `start_cell` [row, col] is the top left cell of this header, Excel cells are 1-indexed."""

    shift_row, shift_col = start_cell[0] - 1, start_cell[1] - 1
    if shift_col >= 26 - 3 - SHOTS_PER_SERIES: # 3 for wireframe
        raise ValueError("shift_col must be less than 26, otherwise the column names will not fit in Excel")

    # date and time
    set_cell(ws, shift_row + 2, shift_col + 2, nowtime(pretty=True), pattern_header, b_left=True, b_right=True, b_top=True, b_bottom=True)
    set_cell(ws, shift_row + 2, shift_col + 3, b_left=True, b_right=True, b_top=True, b_bottom=True) # just border
    ws.merge_cells(start_row=shift_row + 2, start_column=shift_col + 2, end_row=shift_row + 2, end_column=shift_col + 3)

    # legend for colors
    set_cell(ws, shift_row + 2, shift_col + 5, "manuell korrigiert", pattern_mark1, b_left=True, b_right=True, b_top=True, b_bottom=True, center_h=True)
    set_cell(ws, shift_row + 2, shift_col + 6, b_left=True, b_right=True, b_top=True, b_bottom=True) # just border
    ws.merge_cells(start_row=shift_row + 2, start_column=shift_col + 5, end_row=shift_row + 2, end_column=shift_col + 6)
    set_cell(ws, shift_row + 2, shift_col + 8, "Fehlschuss", pattern_mark2, b_left=True, b_right=True, b_top=True, b_bottom=True, center_h=True)
    set_cell(ws, shift_row + 2, shift_col + 9, b_left=True, b_right=True, b_top=True, b_bottom=True) # just border
    ws.merge_cells(start_row=shift_row + 2, start_column=shift_col + 8, end_row=shift_row + 2, end_column=shift_col + 9)

    return (shift_row + 3, shift_col + 10)

def draw_wireframe(ws, shot_data: list[list[Shot]], mode: int, start_cell: tuple[int]=(1, 1)) -> tuple[int]:
    """Draws the wireframe on the excel worksheet. Returns the bottom right cell of this wireframe. \\
    `start_cell` [row, col] is the top left cell of this wireframe, Excel cells are 1-indexed."""

    shift_row, shift_col = start_cell[0] - 1, start_cell[1] - 1
    if shift_col >= 26 - 3 - SHOTS_PER_SERIES: # 3 for wireframe
        raise ValueError("shift_col must be less than 26, otherwise the column names will not fit in Excel")

    # table head, top left
    set_cell(ws, shift_row + 2, shift_col + 2, "Schuss", pattern_header, b_left=True, b_right=True, b_top=True, b_bottom=True)

    for i in range(len(shot_data)):
        # table head, left
        set_cell(ws, shift_row + 3 + i, shift_col + 2, "Ringwert", pattern_header, b_left=True, b_right=True)

        # result column, right
        shot_range = f"{chr(ord('C') + shift_col)}{shift_row + 3 + i}:{chr(ord('C') + shift_col + SHOTS_PER_SERIES - 1)}{shift_row + 3 + i}" # one series
        if mode == 3:
            formula_sum_series = f"=SUMPRODUCT(TRUNC({shot_range}))"
        else:
            formula_sum_series = f"=SUM({shot_range})"
        set_cell(ws, shift_row + 3 + i, shift_col + 3 + SHOTS_PER_SERIES, formula_sum_series, b_left=True, b_right=True) # total sum of series

    # final result, bottom right
    formula_sum_all = f"=SUM({chr(ord('C') + shift_col + SHOTS_PER_SERIES)}{shift_row + 3}:{chr(ord('C') + shift_col + SHOTS_PER_SERIES)}{3 + shift_row + len(shot_data) - 1})"
    set_cell(ws, shift_row + 3 + len(shot_data), shift_col + 3 + SHOTS_PER_SERIES, formula_sum_all, b_left=True, b_right=True, b_top=True, b_bottom=True)

    for i in range(SHOTS_PER_SERIES):
        # table head, top
        set_cell(ws, shift_row + 2, shift_col + 3 + i, i + 1, pattern_header, b_top=True, b_bottom=True, center_h=True)

        # extra border, bottom
        set_cell(ws, shift_row + 3 + len(shot_data), shift_col + 3 + i, b_top=True) # just border

    # extra border, bottom
    set_cell(ws, shift_row + 3 + len(shot_data), shift_col + 2, b_top=True) # just border

    # table head, top right
    set_cell(ws, shift_row + 2, shift_col + 3 + SHOTS_PER_SERIES, "Gesamt", pattern_header, b_left=True, b_right=True, b_top=True, b_bottom=True)

    return (shift_row + 3 + len(shot_data), shift_col + 3 + SHOTS_PER_SERIES + 1)

def fill_wireframe(ws, name_: str, shot_data: list[list[Shot]], mode: int, start_cell: tuple[int]=(1, 1)) -> None:
    """Fills the wireframe in the worksheet with the data. \\
    `start_cell` [row, col] is the top left cell of this wireframe, Excel cells are 1-indexed."""

    shift_row, shift_col = start_cell[0] - 1, start_cell[1] - 1
    if shift_col >= 26 - 3 - SHOTS_PER_SERIES: # 3 for wireframe
        raise ValueError("shift_col must be less than 26, otherwise the column names will not fit in Excel")

    # name, top left outside
    set_cell(ws, shift_row + 2, shift_col + 1, name_)

    # fill data row by row
    for row, series in enumerate(shot_data):
        # name, left outside
        set_cell(ws, shift_row + 3 + row, shift_col + 1, name_)

        # fill data cell by cell
        for col, shot in enumerate(series):
            value = trunc(shot.ring) if mode == 2 else shot.ring
            if shot.ring > 0 and shot.div is None: # manually corrected
                fill = pattern_mark1
            elif shot.ring == 0 and shot.div is None: # missed shot
                fill = pattern_mark2
            else: # normal shot
                fill = None
            set_cell(ws, shift_row + 3 + row, shift_col + 3 + col, value, fill, center_h=True)

def save_data(memory: MemoryHandler, mode: int, start_cell: tuple[int]=(1, 1)) -> str:
    """Saves the data to an Excel file and returns the filepath"""
    wb = openpyxl.Workbook()
    ws = wb.active

    end_cell_current_block = draw_header(ws, start_cell)

    # draw wireframe for each person
    for name, shot_data in memory.MEM_long.items():
        start_cell_new = (end_cell_current_block[0] + 2, start_cell[1]) # leave 2 rows between blocks
        end_cell_current_block = draw_wireframe(ws, shot_data, mode, start_cell=start_cell_new)
        fill_wireframe(ws, name, shot_data, mode, start_cell=start_cell_new)

    # save to file, sorted by year and month

    dir_year = datetime.now().strftime("%Y")
    if not os.path.exists(dir_year):
        os.mkdir(dir_year)

    dir_month = os.path.join(dir_year, datetime.now().strftime("%m"))
    if not os.path.exists(dir_month):
        os.mkdir(dir_month)

    fname = os.path.join(dir_month, f"output_{nowtime()}.xlsx")
    wb.save(fname)
    return str(fname)

def main() -> None:
    """Main function to run the program"""
    global ser

    # sanity check
    if SHOTS_PER_SERIES not in (1, 2, 5, 10) and SHOTS_PER_SERIES % 10 != 0:
        print("Konfigurationsfehler: Schussanzahl pro Serie (SHOTS_PER_SERIES) muss 1, 2, 5, oder ein Vielfaches von 10 sein")
        input("Drücke Enter zum Beenden...")
        raise SystemExit(3)

    # check if the configured serial port exists
    ports_available = [port.name for port in list_ports.comports()]
    if not PORT in ports_available:
        print(f"Fehler: Konfiguriert ist Anschluss {PORT}, wurde nicht gefunden.\n  - bitte Kabelverbindung prüfen\n  - Gerätemanager checken\n  - IT rufen\n\nIm Moment verfügbare Seriellanschlüsse sind:")
        for port in sorted(ports_available):
            print(f"  - {port}")
        input("\nDrücke Enter zum Beenden...")
        raise SystemExit(1)

    # get expected number of shots per strip
    print("Schussanzahl pro Streifen mit Pfeiltasten auswählen und mit Enter bestätigen:")
    SHOTS_PER_STRIP: int = beaupy.select([1, 2, 5, 10], cursor=">", cursor_style="bright_yellow", cursor_index=3)
    print(f"> {Fore.LIGHTCYAN_EX}{SHOTS_PER_STRIP}{Style.RESET_ALL}\n")

    # get processing mode
    modes = [
    "1) mit Teiler",
    "2) ohne Teiler",
    "3) Einzelergebnisse mit Teiler anzeigen, aber ohne Teiler summieren"
    ]
    print("Speicher-Modus mit Pfeiltasten auswählen und mit Enter bestätigen:")
    mode = int(beaupy.select(modes , cursor=">", cursor_style="bright_yellow", return_index=True)) + 1
    print(f"> {Fore.LIGHTCYAN_EX}{modes[mode-1]}{Style.RESET_ALL}\n")

    # setup serial connection and memory handler
    ser = Serial(port=PORT, baudrate=9600, timeout=1, parity=PARITY_NONE, stopbits=STOPBITS_ONE, bytesize=EIGHTBITS, xonxoff=False, rtscts=False)
    mem = MemoryHandler(SHOTS_PER_STRIP)
    ser.write(CODE_NOBAR)
    print("Gerät gefunden -> start")
    sleep(1) # wait so user can read output

    # main logic
    FLAG_save_exit = False # flag to save data and exit program
    while not FLAG_save_exit: # one iteration per person

        # get person name
        clear()
        name_: str = beaupy.prompt(f"Name des {mem.person_count+1}. Schützen eintippen:") # prompt text is cleared after execution
        print(f"Name des {mem.person_count+1}. Schützen eintippen:\n> {Fore.LIGHTCYAN_EX}{name_}{Style.RESET_ALL}\n\n")
        mem.update_person(name_)

        FLAG_next_user = False # flag to jump out of inner while loop to next user
        while not FLAG_next_user: # one iteration per com cycle, one STX per strip

            # start communication cycle
            ser.write(CODE_ENQ) # on ENQ, device always sends NAK or STX, normally never empty
            response = ser.read(1)

            # empty => error
            if response == b"":
                print(f"Keine Antwort vom Gerät erhalten, mögliche Ursachen:\n  - Gerät ist nicht eingeschaltet\n  - Gerät ist nicht angeschlossen\n  - Anschluss {PORT} ist nicht richtig")
                input("Drücke Enter zum Beenden...")
                raise SystemExit(2) # todo overhaul error codes

            # NAK => no new data, check keypresses for exit or next person
            if response == CODE_NAK:
                keys_pressed = record_keypresses(0.5) # detect exit-keypress during 0.5s delay between ENQs, as recommended by manual p. 31
                if keyboard.Key.esc in keys_pressed and keyboard.Key.enter in keys_pressed: # ignore when both pressed
                    continue
                if keyboard.Key.enter in keys_pressed: # enter => new person
                    FLAG_next_user = True # jump out of inner while loop
                    continue
                if keyboard.Key.esc in keys_pressed: # save data and exit
                    FLAG_next_user = True # jump out of inner while loop
                    FLAG_save_exit = True # jump out of outer while loop
                    continue

            # STX => new data being transmitted, start receiving
            if response == CODE_STX: # transmission start
                trans = Transmission.receive(ser)
                if trans is None:
                    print("Fehler: Übertragung fehlerhaft, bitte Kabel auf Wackelkontakt o.ä. prüfen. Dann bestätigen und letzte Scheibe neu erfassen")
                    input("Drücke Enter zum Bestätigen...")
                    continue # ACK is already sent by Transmission.receive()
                mem.update_memory(trans)
                ser.write(CODE_ACK) # com cycle finished

            print(f"Person {mem.person_count} ({Fore.LIGHTYELLOW_EX}{mem.current_person}{Style.RESET_ALL}), Streifen {Fore.LIGHTRED_EX}{mem.strip_count}{Style.RESET_ALL} verarbeitet. {Style.BRIGHT}Weiteren einlegen{Style.RESET_ALL} oder per Tastendruck fortfahren:")
            print(f"    - Drücke {Back.WHITE + Fore.BLACK}<Enter>{Style.RESET_ALL}, um einen neuen Schützen zu erfassen")
            print(f"    - Drücke {Back.WHITE + Fore.BLACK}<Escape>{Style.RESET_ALL}, um die Ergebnisse zu speichern und das Programm zu beenden\n")

    ser.write(CODE_EXIT) # set device inactive
    ser.close()

    mem.finalize()
    if not mem.MEM_long:
        print("Keine Daten zum Speichern vorhanden. (Aus Versehen Escape gedrückt?)")
        input("Drücke Enter zum Beenden...")
        raise SystemExit(4)

    fname = save_data(mem, mode, name_)
    open_file(fname)

if __name__ == "__main__":
    ser: Serial = None
    try:
        main()
    except Exception as e:
        if ser is not None and ser.is_open: # fallback to close serial port gracefully on uncaught error
            ser.write(CODE_EXIT) # set device inactive
            ser.close()
        print(f"nicht abgefangener Fehler aufgetreten: {e}")
        input("Drücke Enter zum Beenden...")

### Terminology in this project ###
# Target : @
# Strip  : |@ @ @ @ @|    (1 thing feeded into the device, contains <SHOTS_PER_STRIP> shots)
# Series : [@ ... @]      (1 row in the output file, contains <SHOTS_PER_SERIES> shots)