from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create a new workbook and get the active sheet
wb = Workbook()
ws = wb.active

# Choose a color (e.g., green)
fill_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

ws.cell(2,3,"test").fill = fill_color

# Save the workbook
wb.save('example.xlsx')
