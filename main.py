import os
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from time import sleep
from math import trunc
from datetime import datetime
from serial import Serial, PARITY_NONE, STOPBITS_ONE, EIGHTBITS

CODE_STX = b"\x02"
CODE_ENQ = b"\x05"
CODE_ACK = b"\x06"
CODE_CR = b"\x0D"
CODE_NAK = b"\x15"
CODE_ETB = b"\x17"
CODE_EXIT = b"\xB0"
CODE_BAR = b"\xB1"
CODE_NOBAR = b"\xB2"
bytemap = {CODE_CR: b"\x22\x3B\x22", # CR -> ";"
           b"\x2E": b"\x2C"} # . -> ,
HEADER = '"Barcode";"Manueller Code";"Scheibentyp";"Anzahl Scheiben";"Teiler-Teilerfaktor";"Anzahl Einschüsse"'

result = []

def clear():
    os.system(("cls" if os.name == "nt" else "clear"))

def nowtime():
    return datetime.now().strftime("%Y_%m_%d-%H_%M_%S")

def truncComma(n: str):
    return int(trunc(float(n.replace(",", "."))))

def saveData(lst: list, mode: str):
    pattern = PatternFill(start_color="00646464", end_color="00646464", fill_type="solid")
    wb = Workbook()
    ws = wb.active
    values = [0]*len(lst)
    for row, line in enumerate(lst, start=1):
        for col, v2 in enumerate(csv.reader([line], delimiter=";", quotechar='"').__next__(), start=1):
            if row != 1 and col >= 7 and col % 4 == 3:
                if mode == "2":
                    ws.cell(row, col, str(truncComma(v2))).fill = pattern
                    values[row] += truncComma(v2)
                else:
                    ws.cell(row, col, v2).fill = pattern
                    values[row] += v2
            else:
                ws.cell(row, col, v2)
    ws.insert_cols(idx=7)
    ws.cell(1, 7, "Gesamt").fill = pattern
    for row, val in enumerate(values, start=2):
        ws.cell(row, 7, val)
    wb.save(f"output_{nowtime()}.xlsx")

def main():
    while True:
        clear()
        print("1) with decimal")
        print("2) truncate")
        mode = input("[1/2] >>> ")
        if mode in ["1", "2"]:
            break
    with Serial(port="/dev/ttyUSB0", baudrate=9600, timeout=1, parity=PARITY_NONE, stopbits=STOPBITS_ONE, bytesize=EIGHTBITS, xonxoff=False, rtscts=False) as ser:
        try:
            ser.write(CODE_NOBAR)
            print("start")
            while True:
                ser.write(CODE_ENQ)
                response = ser.read(1)
                if response == CODE_NAK: # no result
                    sleep(0.5)
                    continue
                if response == CODE_STX: # transmission start
                    data = b"\x22" # "
                    while True:
                        byte = ser.read(1)
                        if byte == CODE_ETB: # end of data
                            break
                        data += bytemap.get(byte, byte)
                    data += b"\x22" # "
                    result.append(data.decode("unicode-escape"))
                    end = None
                    while end != b"\x24":
                        end = ser.read(1)
                    ser.write(CODE_ACK) # com cycle finished
                print("transmission finished")
                sleep(0.5)
        except KeyboardInterrupt:
            print("KeyboardInterrupt")
            ser.write(CODE_EXIT) # set device inactive
            saveData([HEADER]+result, mode)
            sleep(1)

if __name__ == "__main__":
    main()