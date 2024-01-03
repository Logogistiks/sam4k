import csv
import openpyxl
from math import trunc

# Example list of CSV-compatible lines
csv_data = [
    '"Name";"Age";"Location"',
    '"John";"25,3";"New York"',
    '"Alice";"30,5";"Los Angeles";"testabc";"testdef"',
    '"Bob";"28,0";"Chicago"'
]

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
for row, line in enumerate(csv_data, start=1):
    for col, v2 in enumerate(csv.reader([line], delimiter=";", quotechar='"').__next__(), start=1):
        if row != 1 and col == 2:
            ws.cell(row, col, str(trunc(float(v2.replace(",", ".")))))
        else:
            ws.cell(row,)

wb.save("test.xlsx")